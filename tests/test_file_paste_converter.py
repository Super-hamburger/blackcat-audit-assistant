import tempfile
import threading
import time
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from modules.file_paste.converter import (
    ConversionCancelled,
    ConversionControl,
    MARK_ADDRESS_OVERFLOW,
    MARK_QUANTITY_ISSUE,
    NEW_BLACKCAT_HEADERS,
    ONE_PIECE_HEADERS,
    UploadConverter,
)
from modules.file_paste.template import load_upload_template


def color_suffix(fill):
    return (fill.fgColor.rgb or "")[-6:]


class UploadConverterTest(unittest.TestCase):
    def create_workbook(self, path, headers, rows):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])
        workbook.save(path)

    def convert_and_load(self, source_path, temp_dir):
        result = UploadConverter().convert(source_path, Path(temp_dir), open_after=False)
        workbook = load_workbook(result["output_path"])
        return result, workbook.active

    def convert_and_load_one_piece(self, rows):
        with tempfile.TemporaryDirectory() as temp_dir:
            source_path = Path(temp_dir) / "one-piece.xlsx"
            headers = ONE_PIECE_HEADERS + ["数量", "货架", "运输方式"]
            self.create_workbook(source_path, headers, rows)
            result, sheet = self.convert_and_load(source_path, temp_dir)
            values = [[cell.value for cell in row] for row in sheet.iter_rows()]
            fills = {
                coordinate: color_suffix(sheet[coordinate].fill)
                for coordinate in ("AD2", "O2")
            }
        return result, values, fills

    def convert_and_load_blackcat(self, row):
        with tempfile.TemporaryDirectory() as temp_dir:
            source_path = Path(temp_dir) / "blackcat.xlsx"
            headers = NEW_BLACKCAT_HEADERS + ["备注"]
            self.create_workbook(source_path, headers, [row])
            result, sheet = self.convert_and_load(source_path, temp_dir)
            values = [[cell.value for cell in line] for line in sheet.iter_rows()]
            fills = {"B2": color_suffix(sheet["B2"].fill)}
        return result, values, fills

    def one_piece_row(
        self, reference, sku, quantity, shelf, address="芝公園1-2-3", shipping_method="宅急便"
    ):
        return {
            "参考单号": reference,
            "SKU": sku,
            "数量": quantity,
            "货架": shelf,
            "运输方式": shipping_method,
            "收件人": "山田太郎",
            "收件电话": "0312345678",
            "州": "東京都",
            "城市": "港区",
            "地址": address,
            "地址2": "101号室",
            "收件公司": "",
            "收件邮编": "1050011",
        }

    def test_one_piece_sorts_sku_quantity_groups_before_shelf_order(self):
        result, values, _ = self.convert_and_load_one_piece([
            self.one_piece_row("MULTI-LATE", "sku-a,sku-b", "*1+*1", "9-1"),
            self.one_piece_row("ONE-LATE", "sku-c", 1, "8-1"),
            self.one_piece_row("MANY-LATE", "sku-d", 3, "7-1"),
            self.one_piece_row("MULTI-EARLY", "sku-e,sku-f", "*1+*1", "1-1"),
            self.one_piece_row("ONE-EARLY", "sku-g", 1, "2-1"),
            self.one_piece_row("MANY-EARLY", "sku-h", 3, "3-1"),
            self.one_piece_row("MULTI-INVALID", "sku-i,sku-j", 5, "0-1"),
        ])

        self.assertEqual(
            [row[0] for row in values[1:8]],
            [
                "ONE-EARLY",
                "ONE-LATE",
                "MANY-EARLY",
                "MANY-LATE",
                "MULTI-EARLY",
                "MULTI-LATE",
                "MULTI-INVALID",
            ],
        )
        self.assertEqual(result["quantity_issue_count"], 1)

    def test_one_piece_moves_complete_sku_overflow_items_from_ad_to_ac(self):
        sku_one = "a" * 20
        sku_two = "b" * 20
        sku_three = "c" * 20
        result, values, _ = self.convert_and_load_one_piece([
            self.one_piece_row(
                "OVERFLOW",
                f"{sku_one},{sku_two},{sku_three}",
                "*1+*1+*1",
                "1-1",
            ),
        ])

        self.assertEqual(values[1][28], f"{sku_three}*1")
        self.assertEqual(values[1][29], f"{sku_one}*1,{sku_two}*1")
        self.assertLessEqual(len(values[1][29]), 50)
        self.assertEqual(result["quantity_issue_count"], 0)

    def test_one_piece_moves_a_single_overlong_sku_item_to_ac(self):
        long_sku = "x" * 51
        _, values, _ = self.convert_and_load_one_piece([
            self.one_piece_row("LONG-SKU", long_sku, 1, "1-1"),
        ])

        self.assertEqual(values[1][28], f"{long_sku}*1")
        self.assertIsNone(values[1][29])

    def test_blackcat_uses_template_header_sku_and_detail_columns(self):
        result, values, _ = self.convert_and_load_blackcat({
            "单号": "NB-1",
            "收件人电话": "09012345678",
            "收件邮编": "1050011",
            "收件地址": "東京都世田谷区北沢二丁目24-5-101",
            "详细地址": "サンライズマンション 101号室",
            "收件姓名": "王五",
            "发件人电话": "0312345678",
            "sku": "sku-a",
            "明细": "*3",
        })
        template_book, template_sheet = load_upload_template()
        expected_headers = [cell.value for cell in template_sheet[1]]
        template_book.close()

        self.assertEqual(values[0], expected_headers)
        self.assertEqual(len(values[0]), 98)
        self.assertEqual(values[1][0], "NB-1")
        self.assertEqual(values[1][27], "sku-a")
        self.assertEqual(values[1][29], "*3")
        self.assertEqual(result["source_type"], "黑猫新版表格")

    def test_one_piece_writes_zero_for_takkyubin(self):
        _, values, _ = self.convert_and_load_one_piece([
            self.one_piece_row("TAK", "sku-a", 1, "1-1", shipping_method="宅急便"),
        ])

        self.assertEqual(values[1][1], "0")

    def test_blackcat_writes_a_for_toukan_remark(self):
        _, values, fills = self.convert_and_load_blackcat({
            "单号": "TOUKAN",
            "收件人电话": "09012345678",
            "收件邮编": "1050011",
            "收件地址": "東京都港区芝公園1-2-3",
            "详细地址": "101号室",
            "收件姓名": "山田太郎",
            "sku": "sku-a",
            "明细": "*1",
            "备注": "7.15黑猫投函 售后补发订单",
        })

        self.assertEqual(values[1][1], "A")
        self.assertNotEqual(fills["B2"], "F4CCCC")

    def test_blackcat_marks_blank_shipment_type(self):
        _, values, fills = self.convert_and_load_blackcat({
            "单号": "UNKNOWN",
            "收件人电话": "09012345678",
            "收件邮编": "1050011",
            "收件地址": "東京都港区芝公園1-2-3",
            "详细地址": "101号室",
            "收件姓名": "山田太郎",
            "sku": "sku-a",
            "明细": "*1",
            "备注": "",
        })

        self.assertIsNone(values[1][1])
        self.assertEqual(fills["B2"], "F4CCCC")

    def test_converter_marks_ambiguous_quantity_without_dropping_order(self):
        result, values, fills = self.convert_and_load_one_piece([
            self.one_piece_row("AMBIGUOUS", "sku-a,sku-b", 5, "1-1"),
        ])

        self.assertEqual(values[1][0], "AMBIGUOUS")
        self.assertIsNone(values[1][29])
        self.assertEqual(result["quantity_issue_count"], 1)
        self.assertEqual(result["quantity_issue_orders"], ["AMBIGUOUS"])
        self.assertEqual(fills["AD2"], color_suffix(MARK_QUANTITY_ISSUE))

    def test_converter_preserves_overlong_address_in_o_and_marks_it(self):
        long_address = "芝公園一丁目1-2-3 " + "VeryLongBuildingName " * 8 + "101号室"
        result, values, fills = self.convert_and_load_one_piece([
            self.one_piece_row("LONG", "sku-long", 1, "1-1", long_address),
        ])

        self.assertTrue(values[1][14].endswith("101号室"))
        self.assertEqual(result["address_overflow_count"], 1)
        self.assertEqual(fills["O2"], color_suffix(MARK_ADDRESS_OVERFLOW))

    def test_converter_reports_actual_read_and_write_row_progress(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            source_path = Path(temp_dir) / "one-piece.xlsx"
            headers = ONE_PIECE_HEADERS + ["数量", "货架", "运输方式"]
            self.create_workbook(source_path, headers, [
                self.one_piece_row("PROGRESS-1", "sku-a", 1, "1-1"),
                self.one_piece_row("PROGRESS-2", "sku-b", 1, "2-1"),
                self.one_piece_row("PROGRESS-3", "sku-c", 1, "3-1"),
            ])
            events = []

            result = UploadConverter().convert(
                source_path,
                Path(temp_dir) / "output",
                open_after=False,
                progress_callback=events.append,
            )

        reading = [event for event in events if event["phase"] == "reading"]
        writing = [event for event in events if event["phase"] == "writing"]
        self.assertEqual(reading[-1]["current"], 3)
        self.assertEqual(reading[-1]["total"], 3)
        self.assertEqual(writing[-1]["current"], result["row_count"])
        self.assertEqual(writing[-1]["total"], result["row_count"])
        self.assertTrue(any(event["phase"] == "saving" for event in events))

    def test_converter_pauses_between_rows_and_resumes(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            source_path = Path(temp_dir) / "one-piece.xlsx"
            headers = ONE_PIECE_HEADERS + ["数量", "货架", "运输方式"]
            self.create_workbook(source_path, headers, [
                self.one_piece_row("PAUSE-1", "sku-a", 1, "1-1"),
                self.one_piece_row("PAUSE-2", "sku-b", 1, "2-1"),
            ])
            control = ConversionControl()
            first_row_written = threading.Event()
            completed = threading.Event()
            result_box = {}

            def record_progress(event):
                if event["phase"] == "writing" and event["current"] == 1:
                    control.pause()
                    first_row_written.set()

            def convert():
                result_box["result"] = UploadConverter().convert(
                    source_path,
                    Path(temp_dir) / "output",
                    open_after=False,
                    progress_callback=record_progress,
                    control=control,
                )
                completed.set()

            worker = threading.Thread(target=convert)
            worker.start()
            self.assertTrue(first_row_written.wait(3))
            time.sleep(0.1)
            self.assertFalse(completed.is_set())
            control.resume()
            worker.join(5)

        self.assertTrue(completed.is_set())
        self.assertEqual(result_box["result"]["row_count"], 2)

    def test_converter_discards_output_when_cancelled(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            source_path = Path(temp_dir) / "one-piece.xlsx"
            output_dir = Path(temp_dir) / "output"
            headers = ONE_PIECE_HEADERS + ["数量", "货架", "运输方式"]
            self.create_workbook(source_path, headers, [
                self.one_piece_row("CANCEL-1", "sku-a", 1, "1-1"),
                self.one_piece_row("CANCEL-2", "sku-b", 1, "2-1"),
            ])
            control = ConversionControl()

            def cancel_after_first_row(event):
                if event["phase"] == "writing" and event["current"] == 1:
                    control.cancel()

            with self.assertRaises(ConversionCancelled):
                UploadConverter().convert(
                    source_path,
                    output_dir,
                    open_after=False,
                    progress_callback=cancel_after_first_row,
                    control=control,
                )

            self.assertEqual(list(output_dir.glob("*.xlsx")), [])


if __name__ == "__main__":
    unittest.main()
