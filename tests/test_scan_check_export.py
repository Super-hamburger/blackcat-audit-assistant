import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from modules.scan_check.module import ScanCheckService


class ScanCheckExceptionExportTest(unittest.TestCase):
    def service_loaded_with_one_order_and_two_skus(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            source_path = Path(temp_dir) / "source.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["出库单号", "SKU", "数量"])
            sheet.append(["SO-100", "SKU-01", 1])
            sheet.append(["SO-100", "SKU-02", 1])
            workbook.save(source_path)

            service = ScanCheckService()
            service.load_excel(source_path)
            return service

    def test_blocked_scan_is_not_persisted_and_summary_reports_match_counts(self):
        service = self.service_loaded_with_one_order_and_two_skus()
        service.start()
        service.scan("SO-100")
        service.scan("SKU-01")
        logs_before_block = list(service.summary()["recent_logs"])
        blocked = service.scan("UNKNOWN-SKU")

        self.assertEqual(blocked["result"], "block")
        self.assertFalse(hasattr(service, "exception_logs"))
        summary = service.summary()
        self.assertEqual(summary["recent_logs"], logs_before_block)
        self.assertEqual(summary["total_scans"], 2)
        self.assertEqual(summary["failed"], 1)
        self.assertEqual(summary["matched_count"], 1)
        self.assertEqual(summary["matchable_count"], 2)
        self.assertEqual(summary["progress_percent"], 50)

    def test_rescanning_a_completed_known_sku_keeps_its_product_name(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            source_path = Path(temp_dir) / "source.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["出库单号", "SKU", "数量", "商品名称"])
            sheet.append(["SO-100", "SKU-01", 1, "测试商品"])
            workbook.save(source_path)

            service = ScanCheckService()
            service.load_excel(source_path)
            service.start()
            service.scan("SO-100")
            service.scan("SKU-01")
            blocked = service.scan("SKU-01")

            self.assertEqual(blocked["result"], "block")
            self.assertEqual(blocked["product_name"], "测试商品")

    def test_export_unmatched_source_rows_ignores_quantity_two_rows(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            source_path = Path(temp_dir) / "source.xlsx"
            output_path = Path(temp_dir) / "unmatched.csv"
            workbook = Workbook()
            sheet = workbook.active
            headers = ["出库单号", "SKU", "数量", "商品名称", "客户", "备注"]
            matched_source_row = ["SO-100", "SKU-01", 1, "已匹配商品", "客户 A", "保留完整数据"]
            ignored_source_row = ["SO-100", "SKU-02", 2, "忽略商品", "客户 A", "数量为二"]
            unmatched_source_row = ["SO-200", "SKU-03", 1, "未匹配商品", "客户 B", "保留完整数据"]
            sheet.append(headers)
            sheet.append(matched_source_row)
            sheet.append(ignored_source_row)
            sheet.append(unmatched_source_row)
            workbook.save(source_path)

            service = ScanCheckService()
            summary = service.load_excel(source_path)
            self.assertEqual(summary["item_count"], 2)
            self.assertEqual(summary["total_quantity"], 2)

            service.start()
            service.scan("SO-100")
            service.scan("SKU-01")
            output = service.export_unmatched_source_rows(output_path)

            self.assertEqual(Path(output).suffix, ".xlsx")
            exported = load_workbook(output, data_only=True).active
            self.assertEqual(list(exported.iter_rows(values_only=True)), [tuple(headers), tuple(unmatched_source_row)])


if __name__ == "__main__":
    unittest.main()
