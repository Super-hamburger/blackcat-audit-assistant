import sys
import traceback
import json
import tempfile
from pathlib import Path
from PySide6.QtWidgets import QApplication, QMessageBox
from PySide6.QtGui import QIcon
from core.single_instance import SingleInstanceController
from ui.main_window import MainWindow


def runtime_root():
    if hasattr(sys, "_MEIPASS"):
        return Path(sys._MEIPASS)
    return Path(__file__).resolve().parent


def configure_windows_app_id():
    if sys.platform != "win32":
        return
    try:
        import ctypes

        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("blackcat.audit.assistant")
    except Exception:
        pass


def _write_self_test_report(report, output_path):
    if not output_path:
        return
    target = Path(output_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")


def run_self_test(output_path=None):
    report = {"ok": True, "checks": []}

    def record(name, ok, message=""):
        report["checks"].append({"name": name, "ok": bool(ok), "message": str(message)})
        if not ok:
            report["ok"] = False

    try:
        import fitz  # noqa: F401
        record("import fitz", True)
    except Exception as error:
        record("import fitz", False, error)

    try:
        import pymupdf  # noqa: F401
        record("import pymupdf", True)
    except Exception as error:
        record("import pymupdf", False, error)

    try:
        import openpyxl  # noqa: F401
        record("import openpyxl", True)
    except Exception as error:
        record("import openpyxl", False, error)

    try:
        import core.file_safe  # noqa: F401
        import core.utils.file_safe  # noqa: F401
        record("import core file safety modules", True)
    except Exception as error:
        record("import core file safety modules", False, error)

    try:
        from core.sound_engine import SoundEngine

        sound = SoundEngine(enabled=False)
        sound.play_click()
        sound.play_import()
        sound.play_error()
        sound.play_complete()
        sound.play_done()
        record("sound engine compatibility methods", True)
    except Exception as error:
        record("sound engine compatibility methods", False, error)

    try:
        from core.update_installer import UpdateInstaller

        update_guard = UpdateInstaller().prepare_update({})
        record(
            "update installer guarded entry",
            isinstance(update_guard, dict) and update_guard.get("ok") is False and bool(update_guard.get("message")),
            update_guard,
        )
    except Exception as error:
        record("update installer guarded entry", False, error)

    try:
        from core.modules.module_registry import ModuleRegistry

        registry = ModuleRegistry()
        modules = registry.discover()
        module_ids = [module.module_id for module in modules]
        record(
            "discover business modules",
            {"file_paste", "label_compress", "label_printing"}.issubset(module_ids),
            module_ids,
        )

        for module_id in ["file_paste", "label_compress", "label_printing"]:
            try:
                result = registry.run(module_id, {})
                record(f"execute {module_id} entry", result is not None, getattr(result, "message", ""))
            except Exception as error:
                record(f"execute {module_id} entry", False, error)

        try:
            from openpyxl import Workbook

            with tempfile.TemporaryDirectory(prefix="blackcat_self_test_") as temp_name:
                temp_dir = Path(temp_name)
                excel_path = temp_dir / "sample_one_piece.xlsx"
                excel_output_dir = temp_dir / "excel_output"

                wb = Workbook()
                ws = wb.active
                ws.append([
                    "参考单号", "SKU", "运输方式", "收件人", "收件电话",
                    "州", "城市", "地址", "地址2", "收件公司", "收件邮编", "数量", "货架",
                ])
                ws.append([
                    "SELFTEST001", "SKU001", "宅急便", "山田太郎", "0312345678",
                    "東京都", "新宿区", "西新宿1-1-1", "101", "SelfTest", "1600023", 1, "1-1",
                ])
                wb.save(excel_path)

                excel_result = registry.run("file_paste", {
                    "source_path": str(excel_path),
                    "output_dir": excel_output_dir,
                    "open_after": False,
                })
                excel_data = excel_result.data or {}
                excel_output = Path(excel_data.get("output_path", ""))
                record(
                    "run file_paste sample workbook",
                    excel_result.ok
                    and excel_data.get("row_count") == 1
                    and excel_data.get("address_overflow_count") == 0
                    and excel_data.get("quantity_issue_count") == 0
                    and excel_output.exists(),
                    excel_result.message,
                )

                label_source_path = temp_dir / "label_source.xlsx"
                label_finished_path = temp_dir / "label_finished.xlsx"
                label_output_dir = temp_dir / "label_output"

                wb = Workbook()
                ws = wb.active
                ws.append(["客户编号", "参考单号", "SKU", "数量", "货架"])
                ws.append(["12027", "LABELSELFTEST001", "SKU001", 1, "1-1"])
                wb.save(label_source_path)
                wb.close()

                wb = Workbook()
                ws = wb.active
                ws.append([
                    "单号", "收件人电话", "收件邮编", "收件地址", "详细地址", "收件姓名",
                ])
                ws.append([
                    "LABELSELFTEST001", "09012345678", "1000001", "Tokyo", "1-2-3", "Self Test",
                ])
                wb.save(label_finished_path)
                wb.close()

                label_pdf_path = temp_dir / "label_printing_sample.pdf"
                doc = fitz.open()
                page = doc.new_page()
                page.insert_text((72, 72), "TEL 090-1234-5678 a123456789012a")
                doc.save(label_pdf_path)
                doc.close()

                label_result = registry.run("label_printing", {
                    "source_path": str(label_source_path),
                    "finished_path": str(label_finished_path),
                    "pdf_paths": [str(label_pdf_path)],
                    "output_dir": str(label_output_dir),
                    "open_after": False,
                })
                label_data = label_result.data or {}
                label_outputs = [Path(path) for path in label_data.get("output_paths", [])]
                label_page_count = 0
                if len(label_outputs) == 1 and label_outputs[0].exists():
                    output_doc = fitz.open(label_outputs[0])
                    try:
                        label_page_count = output_doc.page_count
                    finally:
                        output_doc.close()
                record(
                    "run label_printing sample pdf",
                    label_result.ok
                    and len(label_outputs) == 1
                    and label_outputs[0].name == "全部客户_合并_货架排序.pdf"
                    and label_page_count == 1,
                    label_result.message,
                )

                pdf_path = temp_dir / "sample_label.pdf"
                pdf_output_dir = temp_dir / "pdf_output"
                doc = fitz.open()
                page = doc.new_page()
                page.insert_text((72, 72), "Tracking number: 123456789012")
                doc.save(pdf_path)
                doc.close()

                pdf_result = registry.run("label_compress", {
                    "pdf_path": str(pdf_path),
                    "output_dir": pdf_output_dir,
                    "batch_size": 90,
                    "auto_zip": False,
                    "open_output": False,
                    "min_run_seconds": 0,
                })
                pdf_data = pdf_result.data or {}
                record(
                    "run label_compress sample pdf",
                    bool(pdf_data.get("success")) and pdf_data.get("saved_count") == 1,
                    pdf_result.message,
                )

                multi_pdf = temp_dir / "sample_multi_candidate_label.pdf"
                multi_output_dir = temp_dir / "multi_pdf_output"
                doc = fitz.open()
                page = doc.new_page(width=327, height=562)
                page.insert_text((50, 80), "TEL 03-1234-5678  ZIP 160-0023")
                page.insert_text((210, 280), "a390123456789a")
                doc.save(multi_pdf)
                doc.close()

                multi_result = registry.run("label_compress", {
                    "pdf_path": str(multi_pdf),
                    "output_dir": multi_output_dir,
                    "batch_size": 90,
                    "auto_zip": False,
                    "open_output": False,
                    "min_run_seconds": 0,
                })
                multi_data = multi_result.data or {}
                multi_output = Path(multi_data.get("output_dir", "")) / "batch_001" / "390123456789.pdf"
                record(
                    "run label_compress multi-candidate pdf",
                    bool(multi_data.get("success")) and multi_output.exists(),
                    multi_result.message,
                )

                low_pdf = temp_dir / "sample_low_confidence_label.pdf"
                low_output_dir = temp_dir / "low_pdf_output"
                doc = fitz.open()
                page = doc.new_page(width=327, height=562)
                page.insert_text((50, 80), "TEL 03-1234-5678")
                page.insert_text((50, 120), "Postal 160-0023")
                doc.save(low_pdf)
                doc.close()

                low_result = registry.run("label_compress", {
                    "pdf_path": str(low_pdf),
                    "output_dir": low_output_dir,
                    "batch_size": 90,
                    "auto_zip": False,
                    "open_output": False,
                    "min_run_seconds": 0,
                })
                low_data = low_result.data or {}
                low_report = Path(low_data.get("recognition_report", ""))
                low_unrecognized = Path(low_data.get("output_dir", "")) / "unrecognized_pages"
                record(
                    "stop label_compress low-confidence pdf",
                    not low_data.get("success") and low_data.get("saved_count") == 0 and low_report.exists() and low_unrecognized.exists(),
                    low_result.message,
                )

                duplicate_pdf = temp_dir / "sample_duplicate_label.pdf"
                duplicate_output_dir = temp_dir / "duplicate_pdf_output"
                doc = fitz.open()
                page = doc.new_page(width=327, height=562)
                page.insert_text((210, 280), "a390123456789a")
                page = doc.new_page(width=327, height=562)
                page.insert_text((210, 280), "a390123456789a")
                doc.save(duplicate_pdf)
                doc.close()

                duplicate_result = registry.run("label_compress", {
                    "pdf_path": str(duplicate_pdf),
                    "output_dir": duplicate_output_dir,
                    "batch_size": 90,
                    "auto_zip": False,
                    "open_output": False,
                    "min_run_seconds": 0,
                })
                duplicate_data = duplicate_result.data or {}
                record(
                    "stop label_compress duplicate pdf",
                    not duplicate_data.get("success") and bool(duplicate_data.get("duplicate_info")),
                    duplicate_result.message,
                )
        except Exception as error:
            record("run sample business workflows", False, error)
    except Exception as error:
        record("module registry", False, error)

    _write_self_test_report(report, output_path)
    return 0 if report["ok"] else 1


def main():
    if "--self-test" in sys.argv:
        output_path = None
        if "--self-test-output" in sys.argv:
            output_index = sys.argv.index("--self-test-output") + 1
            if output_index < len(sys.argv):
                output_path = sys.argv[output_index]
        sys.exit(run_self_test(output_path))

    configure_windows_app_id()
    app = QApplication(sys.argv)
    app.setApplicationName("BlackCat Audit Assistant")
    app.setQuitOnLastWindowClosed(False)

    single_instance = SingleInstanceController("blackcat-audit-assistant")
    if single_instance.is_duplicate:
        return

    icon_path = runtime_root() / "assets" / "icons" / "blackcat_app.ico"
    if icon_path.exists():
        app.setWindowIcon(QIcon(str(icon_path)))
    try:
        window = MainWindow()
        single_instance.activation_requested.connect(window.restore_from_tray)
        window.show()
        sys.exit(app.exec())
    except Exception as error:
        traceback.print_exc()
        QMessageBox.critical(None, "启动失败", f"{error}\n\n{traceback.format_exc()}")
        sys.exit(1)


if __name__ == "__main__":
    main()
