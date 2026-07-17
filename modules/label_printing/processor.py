from dataclasses import dataclass
from datetime import datetime
import csv
from pathlib import Path
import re
from uuid import uuid4

import fitz
from openpyxl import load_workbook

from modules.file_paste.converter import (
    report_progress,
    resolve_sku_quantities,
    shelf_sort_key,
    split_skus,
)


RAW_REQUIRED_HEADERS = ("客户编号", "参考单号", "SKU", "数量", "货架")
FINISHED_HEADER_SCHEMAS = (
    ("单号", "收件人电话", "收件邮编", "收件地址", "详细地址", "收件姓名"),
    (
        "お客様管理番号(内部ID)", "お届け先電話番号", "お届け先郵便番号",
        "お届け先住所", "お届け先住所（アパートマンション名）", "お届け先名",
    ),
)
FINISHED_CANONICAL_HEADERS = (
    "单号", "收件人电话", "收件邮编", "收件地址", "详细地址", "收件姓名",
)


class LabelPrintingError(Exception):
    """Raised when a label page cannot be mapped to one unique order."""


@dataclass(frozen=True)
class PrintOrder:
    reference: str
    customer_id: str
    shelf: str
    sku: str
    quantity: str
    sku_kind: str
    recipient_name: str
    recipient_phone: str
    recipient_postal: str
    recipient_address: str


@dataclass(frozen=True)
class LabelPage:
    pdf_path: Path
    page_index: int
    input_order: int
    label_type: str
    label_id: str
    order: PrintOrder


def normalize_digits(value):
    return "".join(re.findall(r"\d+", str(value or "")))


def normalize_text(value):
    return re.sub(r"[\s\u3000\-－]", "", str(value or "")).casefold()


def classify_sku(sku, quantity):
    item_text, issue = resolve_sku_quantities(sku, quantity)
    if issue is None and len(split_skus(sku)) == 1 and item_text.endswith("*1"):
        return "SKU×1"
    return "SKU×N和多SKU"


def detect_label_type(text):
    normalized = str(text or "")
    return "投函" if "投函" in normalized or "TOUKAN" in normalized.upper() else "宅急便"


def extract_label_id(text):
    match = re.search(r"a(\d{12})a", str(text or ""), re.IGNORECASE)
    if match:
        return match.group(1)
    match = re.search(r"(?<!\d)(\d{12})(?!\d)", str(text or ""))
    return match.group(1) if match else ""


def page_sort_key(page):
    return (*shelf_sort_key(page.order.shelf, page.input_order), page.input_order)


def build_output_groups(pages, scope, split_types):
    pages = [page for page in pages if page.order.sku_kind == "SKU×1"]
    if scope == "all":
        owners = [("全部客户", pages)]
    elif scope == "by_customer":
        grouped = {}
        for page in pages:
            grouped.setdefault(page.order.customer_id, []).append(page)
        owners = [(f"客户{customer_id}", owner_pages) for customer_id, owner_pages in sorted(
            grouped.items(), key=lambda item: shelf_sort_key(item[0], 0)
        )]
    else:
        raise LabelPrintingError(f"不支持的打印范围: {scope}")

    label_types = ("投函", "宅急便") if split_types else ("合并",)
    groups = []
    for owner, owner_pages in owners:
        for label_type in label_types:
            selected = (
                owner_pages if label_type == "合并"
                else [page for page in owner_pages if page.label_type == label_type]
            )
            if selected:
                groups.append((
                    f"{owner}_{label_type}_货架排序.pdf",
                    sorted(selected, key=page_sort_key),
                ))
    return groups


class LabelPrintProcessor:
    def __init__(
        self, source_path, finished_path, pdf_paths, output_dir,
        scope="all", split_types=False, open_after=True,
    ):
        self.source_path = Path(source_path)
        self.finished_path = Path(finished_path)
        self.pdf_paths = [Path(path) for path in pdf_paths]
        self.output_dir = Path(output_dir)
        self.scope = scope
        self.split_types = bool(split_types)
        self.open_after = bool(open_after)
        self._orders = None

    @staticmethod
    def _header_map(sheet, required_headers, workbook_name):
        headers = {
            str(cell.value or "").strip(): cell.column - 1
            for cell in sheet[1]
            if str(cell.value or "").strip()
        }
        missing = [header for header in required_headers if header not in headers]
        if missing:
            raise LabelPrintingError(
                f"{workbook_name}缺少必要列: {', '.join(missing)}"
            )
        return headers

    @classmethod
    def _read_rows(cls, path, required_headers, workbook_name):
        workbook = load_workbook(path, data_only=True, read_only=True)
        try:
            sheet = workbook.worksheets[0]
            headers = cls._header_map(sheet, required_headers, workbook_name)
            rows = []
            for values in sheet.iter_rows(min_row=2, values_only=True):
                rows.append({
                    header: "" if values[index] is None else str(values[index]).strip()
                    for header, index in headers.items()
                })
            return rows
        finally:
            workbook.close()

    @classmethod
    def _read_finished_rows(cls, path):
        workbook = load_workbook(path, data_only=True, read_only=True)
        try:
            sheet = workbook.worksheets[0]
            headers = {
                str(cell.value or "").strip(): cell.column - 1
                for cell in sheet[1]
                if str(cell.value or "").strip()
            }
            schema = next(
                (
                    candidate for candidate in FINISHED_HEADER_SCHEMAS
                    if all(header in headers for header in candidate)
                ),
                None,
            )
            if schema is None:
                expected = " 或 ".join(
                    ", ".join(candidate) for candidate in FINISHED_HEADER_SCHEMAS
                )
                raise LabelPrintingError(f"完整成品表缺少必要列: {expected}")
            rows = []
            for values in sheet.iter_rows(min_row=2, values_only=True):
                rows.append({
                    canonical: "" if values[headers[actual]] is None
                    else str(values[headers[actual]]).strip()
                    for canonical, actual in zip(FINISHED_CANONICAL_HEADERS, schema)
                })
            return rows
        finally:
            workbook.close()

    def load_orders(self):
        source_rows = self._read_rows(
            self.source_path, RAW_REQUIRED_HEADERS, "原始一件代发表"
        )
        source_by_reference = {}
        for row in source_rows:
            reference = row["参考单号"]
            if not reference:
                continue
            if reference in source_by_reference:
                raise LabelPrintingError(f"原始一件代发表参考单号重复: {reference}")
            source_by_reference[reference] = row

        finished_rows = self._read_finished_rows(self.finished_path)
        orders = {}
        for row in finished_rows:
            reference = row["单号"]
            if not reference:
                continue
            if reference in orders:
                raise LabelPrintingError(f"完整成品表参考单号重复: {reference}")
            source = source_by_reference.get(reference)
            if source is None:
                raise LabelPrintingError(f"完整成品表参考单号未在原始表找到: {reference}")
            orders[reference] = PrintOrder(
                reference=reference,
                customer_id=source["客户编号"],
                shelf=source["货架"],
                sku=source["SKU"],
                quantity=source["数量"],
                sku_kind=classify_sku(source["SKU"], source["数量"]),
                recipient_name=row["收件姓名"],
                recipient_phone=row["收件人电话"],
                recipient_postal=row["收件邮编"],
                recipient_address=" ".join(
                    value for value in (row["收件地址"], row["详细地址"]) if value
                ),
            )
        self._orders = orders
        return orders

    @staticmethod
    def _matches_postal_and_name(text, order):
        normalized = normalize_text(text)
        postal = normalize_digits(order.recipient_postal)
        name = normalize_text(order.recipient_name)
        return bool(postal and name and postal in normalize_digits(text) and name in normalized)

    @staticmethod
    def _match_error(page_order, candidates):
        references = ", ".join(order.reference for order in candidates) or "无"
        return LabelPrintingError(
            f"第{page_order + 1}页无法唯一匹配，候选参考单号: {references}"
        )

    def match_page(self, text, page_order):
        orders = self._orders if self._orders is not None else self.load_orders()
        label_type = detect_label_type(text)
        values = list(orders.values())

        if label_type == "投函":
            candidates = [
                order for order in values if self._matches_postal_and_name(text, order)
            ]
        else:
            page_digits = normalize_digits(text)
            phone_candidates = [
                order
                for order in values
                if (phone := normalize_digits(order.recipient_phone)) and phone in page_digits
            ]
            candidates = phone_candidates
            if len(phone_candidates) > 1:
                narrowed_candidates = [
                    order for order in candidates
                    if self._matches_postal_and_name(text, order)
                ]
                candidates = (
                    narrowed_candidates
                    if len(narrowed_candidates) == 1 else phone_candidates
                )

        if len(candidates) != 1:
            raise self._match_error(page_order, candidates)
        return LabelPage(
            pdf_path=Path(),
            page_index=page_order,
            input_order=page_order,
            label_type=label_type,
            label_id="",
            order=candidates[0],
        )

    @staticmethod
    def _task_output_dir(output_dir):
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        return output_dir / f"面单打印_{stamp}_{uuid4().hex[:8]}"

    @staticmethod
    def _write_exception_report(task_dir, problems):
        task_dir.mkdir(parents=True, exist_ok=True)
        report_path = task_dir / "异常报告.csv"
        with report_path.open("w", newline="", encoding="utf-8-sig") as report_file:
            writer = csv.DictWriter(
                report_file,
                fieldnames=("pdf_file", "page", "label_type", "candidates", "reason"),
            )
            writer.writeheader()
            writer.writerows(problems)
        return report_path

    def _read_and_validate_pages(self, documents, progress_callback):
        pages = []
        problems = []
        label_pages = {}
        total_pages = sum(document.page_count for document in documents.values())
        input_order = 0

        for pdf_path, document in documents.items():
            for page_index, pdf_page in enumerate(document):
                text = pdf_page.get_text()
                label_type = detect_label_type(text)
                label_id = extract_label_id(text)
                try:
                    matched = self.match_page(text, input_order)
                except LabelPrintingError as error:
                    reason = str(error)
                    _, separator, candidates = reason.partition("候选参考单号: ")
                    problems.append({
                        "pdf_file": pdf_path.name,
                        "page": page_index + 1,
                        "label_type": label_type,
                        "candidates": candidates if separator else "",
                        "reason": reason,
                    })
                else:
                    page = LabelPage(
                        pdf_path=pdf_path,
                        page_index=page_index,
                        input_order=input_order,
                        label_type=label_type,
                        label_id=label_id,
                        order=matched.order,
                    )
                    pages.append(page)
                    if label_id:
                        duplicate = label_pages.get(label_id)
                        if duplicate is not None:
                            problems.append({
                                "pdf_file": pdf_path.name,
                                "page": page_index + 1,
                                "label_type": label_type,
                                "candidates": ", ".join((
                                    duplicate.order.reference,
                                    page.order.reference,
                                )),
                                "reason": f"重复面单: {label_id}",
                            })
                        else:
                            label_pages[label_id] = page
                input_order += 1
                report_progress(
                    progress_callback,
                    "matching",
                    f"正在匹配面单（{input_order}/{total_pages}）",
                    input_order,
                    total_pages,
                )
        return pages, problems, total_pages

    @staticmethod
    def _remove_created_files(created_paths):
        for path in created_paths:
            try:
                path.unlink(missing_ok=True)
            except OSError:
                pass

    def run(self, progress_callback=None):
        task_dir = self._task_output_dir(self.output_dir)
        documents = {}
        created_paths = []
        try:
            report_progress(
                progress_callback, "opening", "正在读取面单 PDF...", indeterminate=True
            )
            for pdf_path in self.pdf_paths:
                if pdf_path not in documents:
                    documents[pdf_path] = fitz.open(pdf_path)

            pages, problems, total_pages = self._read_and_validate_pages(
                documents, progress_callback
            )
            if problems:
                self._write_exception_report(task_dir, problems)
                raise LabelPrintingError(problems[0]["reason"])

            groups = build_output_groups(pages, self.scope, self.split_types)
            task_dir.mkdir(parents=True, exist_ok=False)
            report_progress(
                progress_callback, "writing", "正在按货架生成打印文件...", 0, len(groups)
            )
            for group_index, (name, group_pages) in enumerate(groups, start=1):
                output_path = task_dir / name
                output_document = fitz.open()
                try:
                    for page in group_pages:
                        output_document.insert_pdf(
                            documents[page.pdf_path],
                            from_page=page.page_index,
                            to_page=page.page_index,
                        )
                    created_paths.append(output_path)
                    output_document.save(output_path)
                finally:
                    output_document.close()
                report_progress(
                    progress_callback,
                    "writing",
                    f"正在生成打印文件（{group_index}/{len(groups)}）",
                    group_index,
                    len(groups),
                )

            report_progress(progress_callback, "completed", "面单打印文件生成完成", 1, 1)
            return {
                "output_dir": str(task_dir),
                "output_paths": [str(path) for path in created_paths],
                "total_pages": total_pages,
                "matched_pages": len(pages),
                "excluded_pages": sum(
                    page.order.sku_kind != "SKU×1" for page in pages
                ),
            }
        except Exception:
            self._remove_created_files(created_paths)
            raise
        finally:
            for document in documents.values():
                document.close()
