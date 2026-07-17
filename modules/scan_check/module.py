from collections import deque
import re
import time
from dataclasses import dataclass, field
from pathlib import Path

from core.modules.module_contract import ModuleResult


def normalize_code(value):
    text = "" if value is None else str(value)
    return re.sub(r"\s+", "", text).upper()


def display_value(value):
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        return text[:-2]
    return text


@dataclass
class ScanItem:
    order_number: str
    sku: str
    product_name: str = ""
    quantity: int = 1
    scanned: int = 0
    row_number: int = 0

    @property
    def status(self):
        if self.scanned >= self.quantity:
            return "已完成"
        if self.scanned > 0:
            return "进行中"
        return "未开始"


@dataclass
class ScanCheckService:
    orders: dict = field(default_factory=dict)
    order_keys: dict = field(default_factory=dict)
    current_order: str = ""
    source_path: str = ""
    loaded: bool = False
    active: bool = False
    paused: bool = False
    total_scans: int = 0
    passed: int = 0
    failed: int = 0
    recent_logs: list = field(default_factory=list)
    source_headers: list = field(default_factory=list)
    eligible_rows: list = field(default_factory=list)
    pending_row_indexes_by_key: dict = field(default_factory=dict)
    matched_row_indexes: set = field(default_factory=set)
    started_at: float = 0.0

    order_aliases = ("出库单号", "出库单", "订单号", "单号", "参考单号", "发货单号", "order", "order no")
    sku_aliases = ("sku", "商品编码", "品番", "条码", "货号", "商品sku", "商品sku编码")
    qty_aliases = ("数量", "出库数量", "发货数量", "qty", "quantity", "件数")
    name_aliases = ("商品名称", "品名", "商品名", "名称", "product", "name")

    def load_excel(self, path):
        from openpyxl import load_workbook

        source = Path(path)
        if not source.exists():
            raise FileNotFoundError(f"Excel 文件不存在：{source}")

        workbook = load_workbook(source, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()
        if not rows:
            raise ValueError("Excel 文件为空。")

        header_index, headers = self._detect_header(rows)
        columns = self._detect_columns(headers)
        if columns["order"] is None or columns["sku"] is None:
            raise ValueError("未识别到出库单号列或 SKU 列，请检查 Excel 表头。")

        orders = {}
        order_keys = {}
        source_headers = list(rows[header_index])
        eligible_rows = []
        pending_row_indexes_by_key = {}
        source_column_count = max(len(source_headers), *(len(row) for row in rows[header_index + 1 :]))
        source_headers.extend([None] * (source_column_count - len(source_headers)))
        for row_offset, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            order = display_value(self._cell(row, columns["order"]))
            sku = display_value(self._cell(row, columns["sku"]))
            if not order or not sku:
                continue

            quantity = self._parse_quantity(self._cell(row, columns["qty"])) if columns["qty"] is not None else 1
            if quantity != 1:
                continue
            name = display_value(self._cell(row, columns["name"])) if columns["name"] is not None else ""
            order_key = normalize_code(order)
            sku_key = normalize_code(sku)
            if not order_key or not sku_key:
                continue

            source_row = list(row) + [None] * (source_column_count - len(row))
            source_row_index = len(eligible_rows)
            eligible_rows.append(source_row)
            pending_row_indexes_by_key.setdefault((order_key, sku_key), deque()).append(source_row_index)

            if order_key not in orders:
                orders[order_key] = {
                    "order_number": order,
                    "items": {},
                    "row_numbers": [],
                }
                order_keys[order_key] = order
            orders[order_key]["row_numbers"].append(row_offset)

            items = orders[order_key]["items"]
            if sku_key not in items:
                items[sku_key] = ScanItem(
                    order_number=order,
                    sku=sku,
                    product_name=name,
                    quantity=quantity,
                    row_number=row_offset,
                )
            else:
                items[sku_key].quantity += quantity
                if not items[sku_key].product_name and name:
                    items[sku_key].product_name = name

        if not orders:
            raise ValueError("Excel 中没有可用的出库单号和 SKU 数据。")

        self.orders = orders
        self.order_keys = order_keys
        self.current_order = ""
        self.source_path = str(source)
        self.loaded = True
        self.active = False
        self.paused = False
        self.total_scans = 0
        self.passed = 0
        self.failed = 0
        self.recent_logs = []
        self.source_headers = source_headers
        self.eligible_rows = eligible_rows
        self.pending_row_indexes_by_key = pending_row_indexes_by_key
        self.matched_row_indexes = set()
        self.started_at = 0.0
        return self.summary()

    def start(self):
        if not self.loaded:
            return self._state("error", "请先导入 Excel。")
        self.active = True
        self.paused = False
        if not self.started_at:
            self.started_at = time.time()
        return self._state("started", "验单中，请扫描出库单号或 SKU。")

    def pause(self):
        if not self.loaded:
            return self._state("error", "请先导入 Excel。")
        self.paused = not self.paused
        if self.paused:
            return self._state("paused", "已暂停扫码验单。")
        self.active = True
        return self._state("resumed", "已继续扫码验单。")

    def scan(self, raw_code):
        code = normalize_code(raw_code)
        shown_code = str(raw_code or "").strip()
        if not self.loaded:
            return self._fail(shown_code, "", "未导入 Excel", count_scan=False)
        if not self.active:
            return self._fail(shown_code, "", "请先点击开始验单", count_scan=False)
        if self.paused:
            return self._fail(shown_code, "", "当前已暂停", count_scan=False)
        if not code:
            return self._fail("", "", "扫码内容为空", count_scan=False)

        if code in self.orders:
            self.current_order = code
            self._append_log(self.orders[code]["order_number"], "", "选中出库单")
            return self._state("order_selected", "已选中出库单。", order_key=code)

        if not self.current_order:
            return self._fail(shown_code, "", "请先扫描出库单号")

        order = self.orders.get(self.current_order)
        if not order:
            self.current_order = ""
            return self._fail(shown_code, "", "当前出库单不存在")

        item = order["items"].get(code)
        if not item:
            return self._fail(order["order_number"], shown_code, "SKU 不属于当前出库单")

        if item.scanned >= item.quantity:
            return self._fail(
                order["order_number"], item.sku, "该 SKU 已扫够数量", product_name=item.product_name
            )

        pending_indexes = self.pending_row_indexes_by_key.get((self.current_order, code))
        if not pending_indexes:
            return self._fail(
                order["order_number"], item.sku, "该 SKU 已扫够数量", product_name=item.product_name
            )

        item.scanned += 1
        self.matched_row_indexes.add(pending_indexes.popleft())
        self.total_scans += 1
        self.passed += 1
        result = self._state("matched", "匹配成功，可以放行。", order_key=self.current_order, sku_key=code)
        result["result"] = "pass"
        self._append_log(order["order_number"], item.sku, "匹配成功")
        return result

    def current_items(self):
        if not self.current_order or self.current_order not in self.orders:
            return []
        return list(self.orders[self.current_order]["items"].values())

    def current_order_number(self):
        if not self.current_order or self.current_order not in self.orders:
            return ""
        return self.orders[self.current_order]["order_number"]

    def progress_percent(self):
        total = len(self.eligible_rows)
        return int(round(len(self.matched_row_indexes) * 100 / total)) if total else 0

    def summary(self):
        order_count = len(self.orders)
        item_count = sum(len(order["items"]) for order in self.orders.values())
        total_quantity = sum(item.quantity for order in self.orders.values() for item in order["items"].values())
        pass_rate = (self.passed / self.total_scans * 100) if self.total_scans else 0
        matched_count = len(self.matched_row_indexes)
        matchable_count = len(self.eligible_rows)
        return {
            "loaded": self.loaded,
            "active": self.active,
            "paused": self.paused,
            "source_path": self.source_path,
            "order_count": order_count,
            "item_count": item_count,
            "total_quantity": total_quantity,
            "current_order": self.current_order_number(),
            "current_order_key": self.current_order,
            "total_scans": self.total_scans,
            "passed": self.passed,
            "failed": self.failed,
            "pass_rate": pass_rate,
            "matched_count": matched_count,
            "matchable_count": matchable_count,
            "progress_percent": self.progress_percent(),
            "unmatched_row_count": matchable_count - matched_count,
            "recent_logs": self.recent_logs[:50],
        }

    def unmatched_source_rows(self):
        return [
            row
            for index, row in enumerate(self.eligible_rows)
            if index not in self.matched_row_indexes
        ]

    def export_unmatched_source_rows(self, output_path):
        output = Path(output_path)
        if output.suffix.lower() != ".xlsx":
            output = output.with_suffix(".xlsx")
        output.parent.mkdir(parents=True, exist_ok=True)

        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "未匹配源数据"
        sheet.append(self.source_headers)
        for row in self.unmatched_source_rows():
            sheet.append(row)
        workbook.save(output)
        workbook.close()
        return str(output)

    def _detect_header(self, rows):
        best_index = 0
        best_score = -1
        best_headers = []
        for index, row in enumerate(rows[:15]):
            headers = [display_value(value) for value in row]
            score = 0
            for header in headers:
                key = normalize_code(header)
                if self._matches(key, self.order_aliases):
                    score += 3
                if self._matches(key, self.sku_aliases):
                    score += 3
                if self._matches(key, self.qty_aliases):
                    score += 1
                if self._matches(key, self.name_aliases):
                    score += 1
            if score > best_score:
                best_index = index
                best_score = score
                best_headers = headers
        return best_index, best_headers

    def _detect_columns(self, headers):
        return {
            "order": self._find_column(headers, self.order_aliases),
            "sku": self._find_column(headers, self.sku_aliases),
            "qty": self._find_column(headers, self.qty_aliases),
            "name": self._find_column(headers, self.name_aliases),
        }

    def _find_column(self, headers, aliases):
        for index, header in enumerate(headers):
            if self._matches(normalize_code(header), aliases):
                return index
        return None

    def _matches(self, header_key, aliases):
        if not header_key:
            return False
        alias_keys = [normalize_code(alias) for alias in aliases]
        return any(alias in header_key or header_key in alias for alias in alias_keys)

    def _cell(self, row, index):
        if index is None or index >= len(row):
            return None
        return row[index]

    def _parse_quantity(self, value):
        try:
            number = int(float(str(value).strip()))
            return max(number, 1)
        except Exception:
            return 1

    def _state(self, status, message, order_key=None, sku_key=None):
        order_key = order_key or self.current_order
        order_number = ""
        sku = ""
        product_name = ""
        if order_key and order_key in self.orders:
            order_number = self.orders[order_key]["order_number"]
            if sku_key and sku_key in self.orders[order_key]["items"]:
                item = self.orders[order_key]["items"][sku_key]
                sku = item.sku
                product_name = item.product_name
        return {
            "result": "neutral",
            "status": status,
            "message": message,
            "order_number": order_number,
            "sku": sku,
            "product_name": product_name,
            "time": time.strftime("%H:%M:%S"),
            "summary": self.summary(),
            "items": [item.__dict__ | {"status": item.status} for item in self.current_items()],
        }

    def _fail(self, order_number, sku, reason, count_scan=True, product_name=""):
        if count_scan:
            self.total_scans += 1
            self.failed += 1
        result = self._state("blocked", reason)
        result["result"] = "block"
        result["order_number"] = order_number or result.get("order_number", "")
        result["sku"] = sku
        result["product_name"] = product_name
        result["message"] = reason
        return result

    def _append_log(self, order_number, sku, result):
        self.recent_logs.insert(0, {
            "time": time.strftime("%H:%M:%S"),
            "order_number": order_number,
            "sku": sku,
            "result": result,
        })
        self.recent_logs = self.recent_logs[:100]

def run(context):
    context = context or {}
    action = context.get("action")
    service = context.get("service")
    if action == "create_service":
        return ModuleResult(ok=True, message="scan_check service created.", data={"service": ScanCheckService()})
    if not isinstance(service, ScanCheckService):
        return ModuleResult(ok=False, message="scan_check module requires a ScanCheckService instance.", data={})
    if action == "load_excel":
        return ModuleResult(ok=True, message="Excel loaded.", data=service.load_excel(context.get("excel_path", "")))
    if action == "start":
        return ModuleResult(ok=True, message="Scan check started.", data=service.start())
    if action == "pause":
        return ModuleResult(ok=True, message="Scan check pause toggled.", data=service.pause())
    if action == "scan":
        return ModuleResult(ok=True, message="Code scanned.", data=service.scan(context.get("code", "")))
    return ModuleResult(
        ok=False,
        message="Unsupported scan_check action.",
        data={"module": "scan_check", "expected_actions": ["create_service", "load_excel", "start", "pause", "scan"]},
    )
