from copy import copy
from datetime import datetime
import os
from pathlib import Path
import re
import unicodedata

from openpyxl.styles import PatternFill

from core.file_safe import ensure_writable_dir
from modules.file_paste.address_splitter import split_japanese_address
from modules.file_paste.template import load_upload_template


ONE_PIECE_HEADERS = [
    "参考单号", "SKU", "收件人", "收件电话", "州", "城市", "地址", "收件邮编",
]

NEW_BLACKCAT_HEADERS = [
    "单号", "收件人电话", "收件邮编", "收件地址", "详细地址",
    "收件姓名", "sku", "明细",
]

MARK_ADDRESS_SPLIT = PatternFill("solid", fgColor="FFF2CC")
MARK_ADDRESS_OVERFLOW = PatternFill("solid", fgColor="F4CCCC")
MARK_QUANTITY_ISSUE = PatternFill("solid", fgColor="F4CCCC")
MARK_MISSING_REQUIRED = PatternFill("solid", fgColor="F4CCCC")

TEXT_COLUMNS = ("A", "I", "K", "L", "M", "N", "O", "P", "AB", "AD")


def norm(value):
    if value is None:
        return ""
    return str(value).strip()


def normalized_header(value):
    return norm(value).replace("\n", "").replace("\r", "").replace(" ", "")


def build_header_map(ws):
    result = {}
    for cell in ws[1]:
        key = normalized_header(cell.value)
        if key:
            result[key] = cell.column
    return result


def has_headers(header_map, names):
    return all(normalized_header(name) in header_map for name in names)


def read(row, header_map, name):
    index = header_map.get(normalized_header(name))
    return "" if index is None else norm(row[index - 1].value)


def split_skus(value):
    return [part.strip() for part in re.split(r"[,，]", norm(value)) if part.strip()]


def resolve_sku_quantities(sku_text, quantity_value):
    skus = split_skus(sku_text)
    if not skus:
        return "", "SKU为空"

    quantity_text = unicodedata.normalize("NFKC", norm(quantity_value))
    explicit = re.fullmatch(r"\*(\d+)(?:\+\*(\d+))*", quantity_text)
    if explicit:
        amounts = [int(amount) for amount in re.findall(r"\*(\d+)", quantity_text)]
        if len(amounts) == len(skus):
            return ",".join(f"{sku}*{amount}" for sku, amount in zip(skus, amounts)), None
        return "", "SKU数量与数量明细项数不一致"

    if re.fullmatch(r"\d+", quantity_text):
        total = int(quantity_text)
        if len(skus) == 1:
            return f"{skus[0]}*{total}", None
        if total == len(skus):
            return ",".join(f"{sku}*1" for sku in skus), None
        return "", "多个SKU只有总数量，无法确认每个SKU数量"

    return "", "数量格式无法识别"


def shelf_sort_key(value, source_index):
    shelf = norm(value)
    if not shelf or len(split_skus(shelf)) != 1:
        return 2, source_index

    parts = re.split(r"[-－]", shelf)
    if all(re.fullmatch(r"[0-9０-９]+", part) for part in parts):
        return 0, tuple(int(unicodedata.normalize("NFKC", part)) for part in parts), source_index
    if re.match(r"^[A-Za-z]", shelf):
        return 1, shelf.casefold(), source_index
    return 2, source_index


def combine_blackcat_address(record):
    parts = [record["address"], record["detail_address"]]
    return " ".join(part for part in parts if part)


def combine_one_piece_address(record):
    main = "".join(record[key] for key in ("prefecture", "city", "street") if record[key])
    return " ".join(part for part in (main, record["apartment"]) if part)


def clone_template_row(sheet, source_row, target_row):
    for column in range(1, sheet.max_column + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        target.value = source.value
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)


def set_text(sheet, coordinate, value):
    cell = sheet[coordinate]
    cell.value = norm(value) if value is not None else None
    cell.number_format = "@"


class UploadConverter:
    def convert(self, source_path, output_dir, open_after=True):
        source_path = Path(source_path)
        output_dir = ensure_writable_dir(output_dir, "output")

        from openpyxl import load_workbook

        source_book = load_workbook(source_path, data_only=True)
        source_sheet = source_book.worksheets[0]
        header_map = build_header_map(source_sheet)

        if has_headers(header_map, ONE_PIECE_HEADERS):
            source_type = "一件代发表格"
            read_row = self.read_one_piece_row
        elif has_headers(header_map, NEW_BLACKCAT_HEADERS):
            source_type = "黑猫新版表格"
            read_row = self.read_new_blackcat_row
        else:
            raise ValueError("无法识别表格格式，需要黑猫新版表或一件代表。")

        records = []
        for source_index, row in enumerate(source_sheet.iter_rows(min_row=2), start=0):
            record = read_row(row, header_map)
            if record["reference"]:
                record["source_index"] = source_index
                records.append(record)

        if source_type == "一件代发表格":
            records.sort(key=lambda record: shelf_sort_key(record["shelf"], record["source_index"]))

        output_book, output_sheet = load_upload_template()
        if output_sheet.max_row > 2:
            output_sheet.delete_rows(3, output_sheet.max_row - 2)

        split_count = 0
        overflow_count = 0
        missing_count = 0
        quantity_issue_orders = []

        for row_offset, record in enumerate(records):
            output_row = row_offset + 2
            if output_row > 2:
                clone_template_row(output_sheet, 2, output_row)

            address_text = (
                combine_one_piece_address(record)
                if source_type == "一件代发表格"
                else combine_blackcat_address(record)
            )
            address = split_japanese_address(address_text)
            if address["was_split"]:
                split_count += 1
            if address["overflow"]:
                overflow_count += 1

            if source_type == "一件代发表格":
                item_text, quantity_issue = resolve_sku_quantities(
                    record["sku"], record["quantity"]
                )
                ab_value = record["shelf"]
            else:
                item_text = record["detail"]
                quantity_issue = None
                ab_value = record["sku"]

            if quantity_issue:
                quantity_issue_orders.append(record["reference"])

            values = {
                "A": record["reference"],
                "E": datetime.now().strftime("%Y%m%d"),
                "I": record["phone"],
                "K": record["postal"],
                "L": address["L"],
                "M": address["M"],
                "N": address["N"],
                "O": address["O"],
                "P": record["recipient"],
                "AB": ab_value,
                "AD": item_text,
            }
            for column, value in values.items():
                set_text(output_sheet, f"{column}{output_row}", value)

            if address["was_split"]:
                for column in ("L", "M", "N", "O"):
                    if address[column]:
                        output_sheet[f"{column}{output_row}"].fill = MARK_ADDRESS_SPLIT
            if address["overflow"]:
                output_sheet[f"O{output_row}"].fill = MARK_ADDRESS_OVERFLOW
            if quantity_issue:
                output_sheet[f"AD{output_row}"].fill = MARK_QUANTITY_ISSUE

            missing_columns = [
                column
                for column, value in (
                    ("I", record["phone"]),
                    ("K", record["postal"]),
                    ("L", address["L"]),
                    ("P", record["recipient"]),
                )
                if not value
            ]
            if missing_columns:
                missing_count += 1
                for column in missing_columns:
                    output_sheet[f"{column}{output_row}"].fill = MARK_MISSING_REQUIRED

        output_path = output_dir / f"黑猫上传表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_book.save(output_path)
        source_book.close()
        output_book.close()

        if open_after:
            os.startfile(output_path)

        return {
            "output_path": str(output_path),
            "row_count": len(records),
            "split_count": split_count,
            "missing_count": missing_count,
            "address_overflow_count": overflow_count,
            "quantity_issue_count": len(quantity_issue_orders),
            "quantity_issue_orders": quantity_issue_orders,
            "source_type": source_type,
        }

    def read_one_piece_row(self, row, header_map):
        return {
            "reference": read(row, header_map, "参考单号"),
            "phone": read(row, header_map, "收件电话"),
            "postal": read(row, header_map, "收件邮编"),
            "recipient": read(row, header_map, "收件人"),
            "prefecture": read(row, header_map, "州"),
            "city": read(row, header_map, "城市"),
            "street": read(row, header_map, "地址"),
            "apartment": read(row, header_map, "地址2"),
            "shelf": read(row, header_map, "货架"),
            "sku": read(row, header_map, "SKU"),
            "quantity": read(row, header_map, "数量"),
        }

    def read_new_blackcat_row(self, row, header_map):
        return {
            "reference": read(row, header_map, "单号"),
            "phone": read(row, header_map, "收件人电话"),
            "postal": read(row, header_map, "收件邮编"),
            "recipient": read(row, header_map, "收件姓名"),
            "address": read(row, header_map, "收件地址"),
            "detail_address": read(row, header_map, "详细地址"),
            "sku": read(row, header_map, "sku"),
            "detail": read(row, header_map, "明细"),
        }
